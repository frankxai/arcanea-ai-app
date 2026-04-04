#!/usr/bin/env python3
"""Generate Arcanea.ai comprehensive sitemap XLSX with 4 sheets."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# Shared styles
hf = Font(bold=True, color='FFFFFF', size=11)
hfill = PatternFill('solid', fgColor='0D47A1')
brd = Border(*(Side(style='thin', color='CCCCCC'),)*4)

def styled_header(ws, headers, widths):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font, cell.fill, cell.border = hf, hfill, brd
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'

# ═══ Sheet 1: Sitemap ═══
ws = wb.active
ws.title = 'Sitemap'
styled_header(ws, ['Route','Page Name','Category','Status','Nav Linked','Vision / Purpose','User Flow','CTA','Design Grade','Visual Assets','Agent Session','Improvement Notes'],
              [28,22,16,10,12,40,30,25,12,20,18,35])

sfill = {'LIVE':PatternFill('solid',fgColor='C8E6C9'),'PARTIAL':PatternFill('solid',fgColor='FFF9C4'),'STUB':PatternFill('solid',fgColor='FFCDD2'),'ARCHIVE':PatternFill('solid',fgColor='EEEEEE'),'EXPERIMENTAL':PatternFill('solid',fgColor='F3E5F5'),'DEV ONLY':PatternFill('solid',fgColor='EEEEEE')}
gfill = {'A-':PatternFill('solid',fgColor='A5D6A7'),'A':PatternFill('solid',fgColor='81C784'),'B+':PatternFill('solid',fgColor='C8E6C9'),'B':PatternFill('solid',fgColor='E8F5E9'),'B-':PatternFill('solid',fgColor='FFF9C4'),'C+':PatternFill('solid',fgColor='FFE0B2'),'D':PatternFill('solid',fgColor='FFCDD2'),'N/A':PatternFill('solid',fgColor='EEEEEE')}

routes = [
    ('/','Homepage','Marketing','LIVE','YES','Landing — first impression, chat box hero','Visit→Chat/Imagine→Signup','Start Creating','B+','Hero ghost portraits','2026-04-03 A4','OG image wired, hero copy fixed'),
    ('/chat','Chat','Create','LIVE','YES','Primary product — AI chat with 16 Luminors','Select agent→Type→Stream','Type a message','B+','ArcaneanMark empty state','2026-04-03 A4','Provider logos added'),
    ('/chat/[luminorId]','Chat Luminor','Create','LIVE','deep link','Direct chat with specific AI','Click luminor→Chat','Continue','B+','Luminor avatar','',''),
    ('/imagine','Imagine','Create','LIVE','YES','AI image generation studio','Pick style→Prompt→Generate 4','Generate','B+','None','2026-04-03 A4','SVGs→Phosphor, copy improved'),
    ('/studio','Studio','Create','PARTIAL','YES','Multi-mode: text, image, code, music','Pick mode→Create→Save','Start creating','B','','','Wire to Supabase'),
    ('/studio/author','Author Studio','Create','PARTIAL','deep link','Writing workspace','Write→AI assist→Save','Start writing','B','','',''),
    ('/studio/image','Image Studio','Create','PARTIAL','deep link','Image generation workspace','Prompt→Generate→Save','Generate','B','','',''),
    ('/worlds','Worlds','Create','LIVE','YES','Build living universes','Create→Add elements→Visualize','Create a World','B','','2026-04-03 AO','World Graph built'),
    ('/agents','Marketplace','Create','LIVE','YES','12 specialist AI agents','Browse→Select→Chat','Try Agent','B+','CosmicParticles','',''),
    ('/agents/[id]','Agent Detail','Create','LIVE','deep link','Agent profile + capabilities','Read→Try→Purchase','Start Chat','B','','',''),
    ('/agents/hub','Skill Tree','Create','LIVE','YES','50 abilities, 10 Gates','Browse→Unlock→Use','Unlock Skill','B','','',''),
    ('/agents/grimoire','Grimoire','Create','PARTIAL','deep link','Agent spellbook','Browse→Learn→Apply','','B-','','',''),
    ('/forge','Forge','Create','LIVE','YES','Creation hub — 4 paths','Pick path→Create→Deploy','Choose path','B+','','2026-04-03 AO','NFT collection added'),
    ('/forge/companion','Companion Forge','Create','LIVE','YES','Create your own AI agent','Name→Configure→Deploy','Create Companion','B+','','','16 archetypes'),
    ('/forge/luminor','Luminor Forge','Create','LIVE','deep link','Create a Luminor','Configure→Test→Launch','Forge Luminor','B','','',''),
    ('/gallery','Gallery','Explore','LIVE','YES','Community creations showcase','Browse→Filter→Like','Create Your Own','B','Showcase fallback','2026-04-03 A4','Needs real imagery'),
    ('/gallery/forge','Gallery Forge','Explore','LIVE','deep link','Forge gallery','Browse forged items','Forge Your Own','B-','','',''),
    ('/gallery/guardians','Guardians Gallery','Explore','LIVE','deep link','Guardian hero portraits','Browse 10 Guardians','Learn More','B+','10 Guardian WebPs','',''),
    ('/discover','Trending','Explore','LIVE','YES','What creators are building','Browse→Follow→Create','Start Creating','B','','',''),
    ('/factions','Factions','Explore','LIVE','YES','8 Origin Classes','Browse→Quiz→Join','Take the Quiz','B+','','2026-03-30','22 lore docs, 202K words'),
    ('/library','Library','Explore','LIVE','YES','200K+ words, 17 collections','Browse→Read→Track','Start Reading','B-','Maylinn OG','2026-04-03 A4','Needs hero section'),
    ('/library/[collection]','Collection','Explore','LIVE','deep link','Library collection','Browse texts→Read','Read','B','','',''),
    ('/library/[collection]/[text]','Text','Explore','LIVE','deep link','Individual text','Read→Mark progress','Continue Reading','B','','',''),
    ('/library/codex','Codex','Explore','LIVE','deep link','Reference codex','Browse→Search→Reference','','B','','',''),
    ('/library/graph','Knowledge Graph','Explore','LIVE','deep link','Library relationship graph','Explore connections','','B','','',''),
    ('/lore','Lore','Explore','LIVE','YES','Mythology & cosmology hub','Explore→Deep dive→Reference','Explore Mythology','B','','','7 sub-components'),
    ('/lore/guardians','Guardians','Explore','LIVE','YES','The Ten Gate-keepers','Browse→Select→Read','Meet Guardians','B+','10 Guardian portraits','',''),
    ('/lore/guardians/[name]','Guardian Detail','Explore','LIVE','deep link','Individual guardian profile','Read story→Explore gate','','B+','Guardian portrait','',''),
    ('/lore/elements','Elements','Explore','LIVE','deep link','Five Elements cosmology','Read→Practice→Apply','Explore Elements','B','','',''),
    ('/lore/gates','Gates','Explore','LIVE','deep link','Ten Gates progression','Learn→Progress→Unlock','Begin the Path','B','','',''),
    ('/lore/godbeasts','Godbeasts','Explore','LIVE','deep link','Divine beasts','Browse→Read origins','Discover Godbeasts','B','Godbeast WebPs v2','',''),
    ('/lore/malachar','Malachar','Explore','LIVE','deep link','The Dark Lord','Read origin story','','B','','',''),
    ('/lore/wisdoms','Wisdoms','Explore','LIVE','deep link','Wisdom teachings','Browse→Contemplate','','B','','',''),
    ('/living-lore','Chronicles','Explore','LIVE','YES','Interactive serialized stories','Read→Choose→Progress','Start Reading','B+','','2026-03-25','7 crew, episodes'),
    ('/living-lore/book','Book','Explore','LIVE','deep link','Reading interface','Read chapters','Continue','B','','',''),
    ('/living-lore/chronicle','Chronicle Hub','Explore','LIVE','deep link','Episode listing','Browse→Read','Read Episode','B','','',''),
    ('/living-lore/crew','Crew','Explore','LIVE','deep link','7-character crew','Browse→Select→Learn','Meet the Crew','B','','',''),
    ('/living-lore/encounter','Encounter','Explore','LIVE','deep link','AI encounters','Interact→Progress','Begin Encounter','B','','',''),
    ('/ecosystem','Ecosystem','Explore','LIVE','YES','27 repos, 43 packages','Browse nodes→Explore','Explore Ecosystem','B+','','2026-04-03 AO','Interactive constellation'),
    ('/developers','Developers','Explore','LIVE','YES','API & SDK documentation','Read→Get keys→Build','Get API Key','B','','',''),
    ('/developers/api','API Docs','Explore','LIVE','deep link','API reference','Read→Implement','','B','','',''),
    ('/research','Research','Explore','LIVE','YES','Intelligence architecture','Explore systems','View Architecture','B','','2026-03-24',''),
    ('/academy','Academy','Learn','LIVE','YES','Ten Gates creative progression','Explore→Courses→Level up','Start Learning','A-','Luminor CDN images','2026-04-03 A4','Strongest page'),
    ('/academy/courses','Courses','Learn','LIVE','YES','5 structured learning paths','Browse→Enroll→Complete','Start Course','B+','','',''),
    ('/academy/courses/[slug]','Course Detail','Learn','LIVE','deep link','Individual course','Learn→Practice→Complete','Continue','B+','','',''),
    ('/academy/houses','Houses','Learn','LIVE','YES','Seven Academy Houses','Explore→Join→Participate','Join House','B','','',''),
    ('/academy/gates','Gates','Learn','LIVE','deep link','Gate system detail','Learn→Progress→Unlock','Open Gate','B','','',''),
    ('/academy/gates/[id]','Gate Detail','Learn','LIVE','deep link','Individual gate','Deep dive→Practice','','B','','',''),
    ('/academy/ranks','Ranks','Learn','PARTIAL','deep link','Magic rank system','View progression','','B-','','',''),
    ('/academy/assessment','Assessment','Learn','PARTIAL','deep link','Skill assessment','Take test→Get result','Start Assessment','B-','','',''),
    ('/academy/certification','Certification','Learn','LIVE','deep link','Agent certification','Apply→Test→Certify','Get Certified','B','','2026-03-30','Dual-track'),
    ('/academy/gate-quiz','Gate Quiz','Learn','PARTIAL','deep link','Gate discovery quiz','Answer→Discover gate','Take Quiz','B','','',''),
    ('/quiz','Origin Quiz','Learn','LIVE','YES','Discover origin class — viral','Answer→Result→Share','Discover Your Class','B+','','','P0 BUILD priority'),
    ('/pricing','Pricing','Marketing','LIVE','YES','Plans and credit system','Compare→Choose→Subscribe','Get Started Free','B+','None','2026-04-03 A4','Needs creation visuals'),
    ('/about','About','Marketing','LIVE','footer','Company and mission','Read→Understand→Join','Start Creating','B','','',''),
    ('/blog','Blog','Marketing','LIVE','footer','Articles and announcements','Browse→Read→Share','Read More','B+','','','6 articles, JSON-LD'),
    ('/blog/[slug]','Blog Post','Marketing','LIVE','deep link','Individual article','Read→Share→Subscribe','Share','B+','','',''),
    ('/contact','Contact','Marketing','LIVE','footer','Contact form','Fill→Submit→Hear back','Send Message','B','','',''),
    ('/faq','FAQ','Marketing','LIVE','footer','Common questions','Search→Read→Resolve','','B','','',''),
    ('/roadmap','Roadmap','Marketing','LIVE','footer','Product roadmap','Browse milestones','Follow Updates','B','','',''),
    ('/privacy','Privacy','Marketing','LIVE','footer','Privacy policy','Read','','C+','','',''),
    ('/terms','Terms','Marketing','LIVE','footer','Terms of service','Read','','C+','','',''),
    ('/welcome','Welcome','Marketing','LIVE','deep link','Welcome page','Read→Explore','Get Started','B','','',''),
    ('/linktree','Linktree','Marketing','LIVE','deep link','Link aggregator','Click links','','B-','','',''),
    ('/auth/login','Login','Auth','LIVE','auth flow','Authentication','Email/OAuth→Verify','Sign In','B','','',''),
    ('/auth/signup','Sign Up','Auth','LIVE','auth flow','Registration','Fill→Verify→Onboard','Create Account','B','','',''),
    ('/auth/forgot-password','Forgot Password','Auth','LIVE','auth flow','Password reset','Enter email→Check inbox','Reset','B','','',''),
    ('/onboarding','Onboarding','Auth','LIVE','redirect','5-step wizard','Welcome→Type→Luminor→Create','Continue','B+','','','90% complete'),
    ('/dashboard','Dashboard','Auth','LIVE','user menu','User stats dashboard','View→Quick actions→Create','Start Creating','B','','',''),
    ('/dashboard/analytics','Analytics','Auth','LIVE','deep link','Usage analytics','View data→Optimize','','B-','','',''),
    ('/profile','Profile','Auth','LIVE','user menu','User profile hub','View→Edit→Share','Edit Profile','B','','',''),
    ('/profile/[username]','Public Profile','Auth','LIVE','deep link','Public user page','View→Follow→Connect','Follow','B','','',''),
    ('/profile/edit','Edit Profile','Auth','LIVE','deep link','Profile editor','Edit→Save','Save Changes','B','','',''),
    ('/settings','Settings','Auth','LIVE','user menu','User preferences','Configure→Save','Save','B-','','','Needs Supabase'),
    ('/settings/memory','Memory Settings','Auth','LIVE','deep link','AI memory config','Configure→Save','','B-','','',''),
    ('/settings/providers','Providers','Auth','LIVE','deep link','AI provider config','Add keys→Select default','Save','B-','','',''),
    ('/community','Community','Community','LIVE','explore','Community hub','Browse→Join→Engage','Join Discord','B','','',''),
    ('/luminors','Luminors','Community','LIVE','deep link','16 Luminor directory','Browse→Select→Chat','Chat','B+','20 Luminor assets','',''),
    ('/luminors/[id]','Luminor Detail','Community','LIVE','deep link','Individual Luminor','Read→Chat→Bond','Start Chat','B','','',''),
    ('/companions','Companions','Community','LIVE','deep link','AI companion system','Browse→Select→Bond','Meet Companions','B','','',''),
    ('/companions/forge','Companion Forge','Community','LIVE','deep link','Build companion','Configure→Create','Create','B+','','',''),
    ('/council','Council','Community','LIVE','deep link','Luminor council','Convene→Discuss→Decide','Convene','B','','',''),
    ('/council/convening','Convening','Community','LIVE','deep link','Council session','Multi-agent discussion','','B','','',''),
    ('/leaderboard','Leaderboard','Community','LIVE','deep link','Creator rankings','View→Compete→Rise','View Rankings','B-','','',''),
    ('/challenges','Challenges','Community','LIVE','deep link','Creative challenges','Browse→Enter→Submit','Accept Challenge','B-','','',''),
    ('/activity','Activity','Community','LIVE','deep link','Activity feed','Browse updates','','B-','','',''),
    ('/feedback','Feedback','Community','LIVE','deep link','User feedback','Submit→Track','Send Feedback','B-','','',''),
    ('/creations','Creations','Creator Tools','LIVE','user menu','Saved work gallery','Browse→Edit→Share','Create New','B','','2026-03-30',''),
    ('/prompt-books','Prompt Books','Creator Tools','PARTIAL','deep link','Prompt collections','Browse→Create→Organize','Create Collection','B','','','Local only'),
    ('/workspace','Workspace','Creator Tools','STUB','deep link','Creator workspace','Tabbed UI','','D','','','Non-functional'),
    ('/world-builder','World Builder','Creator Tools','STUB','deep link','World building tool','Design→Build→Publish','Start Building','D','','','Needs backend'),
    ('/universe-builder','Universe Builder','Creator Tools','STUB','deep link','Universe builder','Design→Build','','D','','','Needs backend'),
    ('/character-book','Character Book','Creator Tools','PARTIAL','deep link','Character reference','Browse→Create→Reference','Create Character','B-','','',''),
    ('/vision-board','Vision Board','Creator Tools','STUB','deep link','Creative vision board','Plan→Organize→Execute','','D','','','Needs backend'),
    ('/docs','Docs','Docs','LIVE','YES','Documentation hub','Browse→Read→Implement','Get Started','B','','',''),
    ('/docs/api','API Docs','Docs','LIVE','deep link','API reference','Read→Implement','','B','','',''),
    ('/docs/mcp','MCP Docs','Docs','LIVE','deep link','MCP server docs','Install→Configure→Use','Install MCP','B+','','2026-04-03 AO',''),
    ('/docs/mcp/install','MCP Install','Docs','LIVE','deep link','Installation guide','Follow steps','','B','','',''),
    ('/docs/mcp/tools','MCP Tools','Docs','LIVE','deep link','42 tool reference','Browse→Learn→Use','','B','','2026-04-04 AO','42 tools documented'),
    ('/changelog','Changelog','Docs','LIVE','deep link','Version history','Browse updates','','B-','','',''),
    ('/status','Status','Docs','LIVE','deep link','System status','Check→Report','','B-','','',''),
    ('/standards','Standards','Docs','LIVE','deep link','Technical standards','Read→Apply','','B-','','',''),
    ('/architecture','Architecture','Docs','LIVE','deep link','System architecture','Explore tabs','','B','','2026-03-24','ReactFlow diagram'),
    ('/ops','Ops Center','Internal','LIVE','NO','Operations dashboard','Monitor→Debug','','B','','2026-03-30','Auth-gated'),
    ('/command-center','Command Center','Internal','LIVE','NO','Central command','Manage→Deploy','','B-','','','Auth-gated'),
    ('/command','Command','Internal','LIVE','NO','Command hub','Execute→Monitor','','B-','','',''),
    ('/workflows','Workflows','Internal','LIVE','NO','Automation','Design→Run','','B-','','',''),
    ('/music','Music','Content','LIVE','navbar?','14 Suno tracks','Browse→Listen→Download','Listen Now','B','','2026-03-12','Suno embeds'),
    ('/books','Books','Content','LIVE','deep link','Book browser','Browse→Read→Track','Start Reading','B','','',''),
    ('/saga','Saga','Content','LIVE','deep link','Saga hub','Browse→Read→Notes','Read Saga','B','','',''),
    ('/records','Records','Content','LIVE','deep link','Music studio hub','Browse→Listen','Listen','B-','','',''),
    ('/claw','ArcaneaClaw','Content','LIVE','deep link','Media engine','Browse media','','B-','','',''),
    ('/v1','V1','Archive','ARCHIVE','NO','Version 1 archive','','','N/A','','',''),
    ('/v2','V2','Archive','ARCHIVE','NO','Version 2 archive','','','N/A','','',''),
    ('/v3','V3','Archive','ARCHIVE','NO','Version 3 archive','','','N/A','','',''),
    ('/v4','V4','Archive','ARCHIVE','NO','Version 4 archive','','','N/A','','',''),
    ('/chess','Chess','Archive','EXPERIMENTAL','NO','Chess game','','','N/A','','',''),
    ('/components','Components','Internal','DEV ONLY','NO','Component showcase','','','N/A','','',''),
]

for i, row in enumerate(routes, 2):
    for j, val in enumerate(row, 1):
        cell = ws.cell(row=i, column=j, value=val)
        cell.border = brd
        cell.alignment = Alignment(vertical='top', wrap_text=True)
    s = ws.cell(row=i, column=4)
    if s.value in sfill: s.fill = sfill[s.value]
    g = ws.cell(row=i, column=9)
    if g.value in gfill: g.fill = gfill[g.value]
    ws.row_dimensions[i].height = 26

ws.auto_filter.ref = f'A1:L{len(routes)+1}'

# ═══ Sheet 2: Dashboard ═══
ds = wb.create_sheet('Dashboard')
ds['A1'] = 'Arcanea.ai Sitemap Dashboard — April 4, 2026'
ds['A1'].font = Font(bold=True, size=16, color='0D47A1')
ds.merge_cells('A1:D1')
for r, (k, v) in enumerate([
    ('Total Page Routes', len(routes)),('Nav-Linked', 25),('LIVE', sum(1 for r in routes if r[3]=='LIVE')),
    ('PARTIAL', sum(1 for r in routes if r[3]=='PARTIAL')),('STUB', sum(1 for r in routes if r[3]=='STUB')),
    ('ARCHIVE', sum(1 for r in routes if r[3] in ('ARCHIVE','EXPERIMENTAL','DEV ONLY'))),
    ('',''),('Design Grades',''),
    ('A- or better', sum(1 for r in routes if r[8] in ('A-','A','A+'))),
    ('B+ to B', sum(1 for r in routes if r[8] in ('B+','B'))),
    ('B- or lower', sum(1 for r in routes if r[8] in ('B-','C+','C','D'))),
    ('',''),('Infrastructure',''),('Vercel Project','arcanea-ai-appx'),('Team','starlight-intelligence'),
    ('Domain','arcanea.ai'),('Framework','Next.js 16 + React 19'),('Deploy Status','READY'),
    ('Root Directory','apps/web'),
], 3):
    ds.cell(row=r, column=1, value=k).font = Font(bold=True) if not v else Font()
    ds.cell(row=r, column=2, value=v)
ds.column_dimensions['A'].width = 22
ds.column_dimensions['B'].width = 30

# ═══ Sheet 3: Visual Assets ═══
va = wb.create_sheet('Visual Assets')
styled_header(va, ['Directory','Count','Format','Usage','Notes'], [28,10,15,30,40])
for i, row in enumerate([
    ('/public/brand/',4,'JPG+SVG','Logo, hero, OG','OG now wired'),
    ('/public/guardians/',10,'WebP','Hero portraits v1','Lore, academy'),
    ('/public/guardians/gallery/',40,'WebP','4 per guardian','Mostly unused'),
    ('/public/guardians/v2/',18,'WebP','Divine Bonds+Godbeasts','Lore pages'),
    ('/public/guardians/v3/',10,'WebP','Hero v3','OG images'),
    ('/public/guardians/v4/',10,'WebP','Hero v4','Latest version'),
    ('/public/images/luminors/',56,'JPG+WebP','20 Luminors+cover','Luminors page'),
    ('/public/images/forge/',4,'PNG','Ships/vehicles','Forge page'),
    ('/public/icons/tabler/',50,'SVG','Icon subset','Mostly @phosphor'),
    ('TOTAL',188,'Mixed','','~70% unused'),
], 2):
    for j, val in enumerate(row, 1):
        va.cell(row=i, column=j, value=val).border = brd

# ═══ Sheet 4: Action Items ═══
ai_sheet = wb.create_sheet('Action Items')
styled_header(ai_sheet, ['Priority','Action','Area','Impact','Status','Owner'], [10,40,18,25,12,15])
ai_sheet['A1'].fill = PatternFill('solid', fgColor='D32F2F')
for c in range(2,7): ai_sheet.cell(row=1, column=c).fill = PatternFill('solid', fgColor='D32F2F')
pfill = {'P0':PatternFill('solid',fgColor='FFCDD2'),'P1':PatternFill('solid',fgColor='FFF9C4'),'P2':PatternFill('solid',fgColor='E3F2FD'),'P3':PatternFill('solid',fgColor='EEEEEE')}
for i, row in enumerate([
    ('P0','Supabase OAuth config','/auth','Blocks all auth','BLOCKED','Frank'),
    ('P0','Verify OG image on shares','All','Social cards','DONE','A4'),
    ('P0','Origin Class Quiz viral','/quiz','Conversion','TODO',''),
    ('P1','Real images in Gallery','/gallery','Visual credibility','TODO',''),
    ('P1','Creation visuals on Pricing','/pricing','Conversion','TODO',''),
    ('P1','Newsletter form API','Footer','Lead capture','TODO',''),
    ('P1','Library hero section','/library','First impression','TODO',''),
    ('P1','npm publish 13 packages','OSS','Distribution','BLOCKED','npm login'),
    ('P2','Prune ~100 stub pages','All','Perf+clarity','TODO',''),
    ('P2','Unify teal token','Design system','Consistency','TODO',''),
    ('P2','Core Web Vitals audit','All','SEO+perf','TODO',''),
    ('P3','Full Phosphor icon migration','Multiple','Consistency','TODO',''),
], 2):
    for j, val in enumerate(row, 1):
        cell = ai_sheet.cell(row=i, column=j, value=val)
        cell.border = brd
        if j == 1 and val in pfill: cell.fill = pfill[val]

wb.save('C:/Users/frank/Arcanea/docs/ARCANEA_SITEMAP_2026-04-04.xlsx')
print(f'Created: docs/ARCANEA_SITEMAP_2026-04-04.xlsx ({len(routes)} routes, 4 sheets)')
